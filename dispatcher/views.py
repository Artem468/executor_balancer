import datetime
import io
from datetime import date

from django.http import HttpResponse
from drf_spectacular.utils import extend_schema, OpenApiParameter
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from dispatcher.serializer import DispatchSerializer, DailySummaryQuerySerializer
from .models import DispatchLogs
from .tasks import dispatch_request


class DispatcherView(APIView):
    @extend_schema(
        tags=["Распределение"],
        summary="Распределитель",
        description="Получает данные заявки для распределения, решает какой юзер более релевантен",
        request=DispatchSerializer,
        responses={200: {"status": "string", "message": "string"}},
    )
    def post(self, request):
        serializer = DispatchSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        task = dispatch_request.delay(request.data)
        return Response({"task_id": task.id}, status=status.HTTP_202_ACCEPTED)


class DailySummaryView(APIView):
    """
    API вью для получения суммарной выгрузки заявок по дням.
    Поддерживает query-параметры: start_date и end_date в формате YYYY-MM-DD.
    """

    @extend_schema(
        tags=["Распределение"],
        parameters=[
            OpenApiParameter(
                name="start_date",
                type=str,
                description="Начальная дата в формате YYYY-MM-DD (опционально)",
                required=False,
            ),
            OpenApiParameter(
                name="end_date",
                type=str,
                description="Конечная дата в формате YYYY-MM-DD (опционально)",
                required=False,
            ),
        ],
        responses={
            200: {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "date": {"type": "string", "format": "date"},
                        "count": {"type": "integer"},
                    },
                },
                "description": "Список с суммой заявок по дням",
            },
            400: {
                "type": "object",
                "properties": {
                    "error": {"type": "string"},
                },
                "description": "Ошибка в формате даты",
            },
        },
        summary="Получить суммарную выгрузку заявок по дням",
        description="Возвращает количество заявок, сгруппированных по дням. Можно фильтровать по диапазону дат.",
    )
    def get(self, request):
        serializer = DailySummaryQuerySerializer(data=request.query_params)
        if not serializer.is_valid():
            return Response(serializer.errors, status=400)

        start_date = serializer.validated_data.get("start_date")
        end_date = serializer.validated_data.get("end_date")

        summary = DispatchLogs.daily_summary(start_date=start_date, end_date=end_date)
        return Response(summary)


from datetime import date

from drf_spectacular.utils import extend_schema, OpenApiParameter
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from dispatcher.serializer import DispatchSerializer, DailySummaryQuerySerializer
from .models import DispatchLogs
from .tasks import dispatch_request


class DispatcherView(APIView):
    @extend_schema(
        tags=["Распределение"],
        summary="Распределитель",
        description="Получает данные заявки для распределения, решает какой юзер более релевантен",
        request=DispatchSerializer,
        responses={200: {"status": "string", "message": "string"}},
    )
    def post(self, request):
        serializer = DispatchSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        task = dispatch_request.delay(request.data)
        return Response({"task_id": task.id}, status=status.HTTP_202_ACCEPTED)


class DailySummaryView(APIView):
    """
    API вью для получения суммарной выгрузки заявок по дням.
    Поддерживает query-параметры: start_date и end_date в формате YYYY-MM-DD.
    """

    @extend_schema(
        tags=["Распределение"],
        parameters=[
            OpenApiParameter(
                name="start_date",
                type=str,
                description="Начальная дата в формате YYYY-MM-DD (опционально)",
                required=False,
            ),
            OpenApiParameter(
                name="end_date",
                type=str,
                description="Конечная дата в формате YYYY-MM-DD (опционально)",
                required=False,
            ),
        ],
        responses={
            200: {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "date": {"type": "string", "format": "date"},
                        "count": {"type": "integer"},
                    },
                },
                "description": "Список с суммой заявок по дням",
            },
            400: {
                "type": "object",
                "properties": {
                    "error": {"type": "string"},
                },
                "description": "Ошибка в формате даты",
            },
        },
        summary="Получить суммарную выгрузку заявок по дням",
        description="Возвращает количество заявок, сгруппированных по дням. Можно фильтровать по диапазону дат.",
    )
    def get(self, request):
        serializer = DailySummaryQuerySerializer(data=request.query_params)
        if not serializer.is_valid():
            return Response(serializer.errors, status=400)

        start_date = serializer.validated_data.get("start_date")
        end_date = serializer.validated_data.get("end_date")

        summary = DispatchLogs.daily_summary(start_date=start_date, end_date=end_date)
        return Response(summary)


def _parse_date_param(value):
    """
    Парсит YYYY-MM-DD -> datetime.date. Возвращает None если value пустой.
    Бросает ValueError если неверный формат.
    """
    if not value:
        return None
    return datetime.datetime.strptime(value, "%Y-%m-%d").date()


@extend_schema(
    tags=["Логи"],
    parameters=[
        OpenApiParameter(
            name="start_date",
            description="Дата начала (включительно) в формате YYYY-MM-DD",
            required=False,
            type=str,
        ),
        OpenApiParameter(
            name="end_date",
            description="Дата конца (включительно) в формате YYYY-MM-DD",
            required=False,
            type=str,
        ),
    ],
    summary="Экспорт ежедневной сводки логов в Excel",
    description="Возвращает xlsx-файл со сводкой количества заявок по дням. "
                "Параметры start_date и end_date принимаются в формате YYYY-MM-DD.",
)
class ExportDispatchSummaryExcelView(APIView):
    authentication_classes = []
    permission_classes = []

    def get(self, request, *args, **kwargs):
        start_date_str = request.GET.get("start_date")
        end_date_str = request.GET.get("end_date")

        try:
            start_date = _parse_date_param(start_date_str)
            end_date = _parse_date_param(end_date_str)
        except ValueError:
            return Response(
                {"error": "Неверный формат даты. Используйте YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        summary = DispatchLogs.daily_summary(start_date=start_date, end_date=end_date)

        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Summary"

        headers = ["Дата", "Количество заявок"]
        ws.append(headers)

        total = 0
        for row in summary:
            ws.append([row.get("date"), row.get("count", 0)])
            total += int(row.get("count", 0))

        ws.append([])
        ws.append(["Итого", total])

        for i, column_cells in enumerate(ws.columns, 1):
            length = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in column_cells
            )
            ws.column_dimensions[get_column_letter(i)].width = min(50, length + 2)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        sd = start_date_str or "from_begin"
        ed = end_date_str or "to_now"
        filename = f"dispatch_summary_{sd}_{ed}.xlsx"

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
